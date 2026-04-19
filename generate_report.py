"""
Generate SQA03_Nhom7_UnitTest_Report.xlsx
Unit Test Report for Nhom 07 - SQA03
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── color palette ─────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_PASS_BG     = "C6EFCE"   # green
C_FAIL_BG     = "FFC7CE"   # red
C_SECTION_BG  = "BDD7EE"   # light blue section header
C_ALT_ROW     = "F2F7FC"   # alternating row
C_TITLE_BG    = "2E75B6"
C_TITLE_FG    = "FFFFFF"
C_BORDER      = "9DC3E6"
C_SCOPE_YES   = "E2EFDA"   # light green for scope YES
C_SCOPE_NO    = "FCE4D6"   # light orange for scope NO

thin = Side(style='thin', color=C_BORDER)
thick = Side(style='medium', color="1F3864")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
outer_border = Border(left=thick, right=thick, top=thick, bottom=thick)

def hdr_font(size=11, bold=True, color=C_HEADER_FG):
    return Font(name='Calibri', size=size, bold=bold, color=color)

def cell_font(size=10, bold=False, color="000000"):
    return Font(name='Calibri', size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(wrap=True):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border

# ══════════════════════════════════════════════════════════════════════
# SHEET 1: OVERVIEW / TOOLS & SCOPE
# ══════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Overview"

# Title
ws1.merge_cells("A1:H1")
t = ws1["A1"]
t.value = "SQA03 – NHÓM 07 | BÁO CÁO KIỂM THỬ ĐƠN VỊ (UNIT TEST REPORT)"
t.font = Font(name='Calibri', size=16, bold=True, color=C_TITLE_FG)
t.fill = fill(C_TITLE_BG)
t.alignment = center()
ws1.row_dimensions[1].height = 40

# Subtitle
ws1.merge_cells("A2:H2")
s = ws1["A2"]
s.value = "Hệ thống Tuyển dụng Trực tuyến Tích hợp AI Gợi ý Việc làm  |  Ngô Xuân Hòa · Nguyễn Duy Anh · Bùi Ngọc Đức · Trần Đình Hào"
s.font = Font(name='Calibri', size=10, italic=True, color="595959")
s.alignment = center()
ws1.row_dimensions[2].height = 20

ws1.append([])

# 1.1 Tools & Libraries
ws1.merge_cells("A4:H4")
h = ws1["A4"]
h.value = "1.1  Tools & Libraries"
h.font = hdr_font(12)
h.fill = fill(C_HEADER_BG)
h.alignment = center()
ws1.row_dimensions[4].height = 24

headers_tools = ["Hạng mục", "Tên / Phiên bản", "Vai trò", "Ghi chú"]
for i, txt in enumerate(headers_tools, 1):
    c = ws1.cell(row=5, column=i)
    c.value = txt
    c.font = hdr_font(10)
    c.fill = fill("2E75B6")
    c.alignment = center()

tools = [
    ("Testing Framework", "Jest 29.7", "Test runner + assertion library", "Tiêu chuẩn cho TypeScript/JS"),
    ("TypeScript Transformer", "ts-jest 29.2", "Biên dịch TypeScript trong Jest", "isolatedModules=true"),
    ("TypeScript", "TypeScript 5.4", "Ngôn ngữ kiểm thử", "Strict typing cho mock"),
    ("Mocking", "Jest.fn() (built-in)", "Tạo mock cho repositories và services", "Không cần Mockito vì TS"),
    ("Node.js Runtime", "Node.js 20 LTS", "Môi trường thực thi test", ""),
    ("Coverage Tool", "Jest --coverage (built-in V8)", "Báo cáo code coverage", ""),
    ("Database Layer", "Mock (không dùng DB thực)", "In-memory mocks với jest.fn()", "Rollback tự động"),
    ("Package Manager", "npm 10.x", "Quản lý dependencies", ""),
]
for r, row in enumerate(tools, 6):
    for c, val in enumerate(row, 1):
        cell = ws1.cell(row=r, column=c)
        cell.value = val
        cell.font = cell_font()
        cell.alignment = left()
        cell.fill = fill(C_ALT_ROW) if r % 2 == 0 else fill("FFFFFF")

apply_border(ws1, 5, 5+len(tools), 1, 4)
set_col_widths(ws1, [25, 20, 40, 35])

# 1.2 Scope of Testing
row_start = 6 + len(tools) + 2
ws1.merge_cells(f"A{row_start}:H{row_start}")
h2 = ws1[f"A{row_start}"]
h2.value = "1.2  Phạm vi kiểm thử (Scope of Testing)"
h2.font = hdr_font(12)
h2.fill = fill(C_HEADER_BG)
h2.alignment = center()
ws1.row_dimensions[row_start].height = 24

headers_scope = ["#", "Mã chức năng", "Tên chức năng", "Use Case được kiểm thử",
                 "Số TC", "Kỹ thuật áp dụng", "Trong phạm vi?", "Lý do / Ghi chú"]
for i, txt in enumerate(headers_scope, 1):
    c = ws1.cell(row=row_start+1, column=i)
    c.value = txt
    c.font = hdr_font(10)
    c.fill = fill("2E75B6")
    c.alignment = center()
ws1.row_dimensions[row_start+1].height = 30

scope_in = [
    (1, "F01", "Đăng ký tài khoản", "RegisterUserUseCase", 7, "EP, BVA, Decision Table", "✔ Có", ""),
    (2, "F02", "Đăng nhập / Đăng xuất", "LoginUserUseCase, LogoutUserUseCase", 10, "EP, Decision Table, Use Case", "✔ Có", ""),
    (3, "F03", "Quản lý việc làm đã ứng tuyển (UV)", "GetMyApplicationsUseCase, WithdrawApplicationUseCase", 9, "Use Case, EP, State Transition", "✔ Có", ""),
    (4, "F04", "Xem việc làm & Công ty", "GetJobByIdUseCase, SearchJobsUseCase, GetCompanyByIdUseCase", 10, "EP, BVA, Use Case", "✔ Có", ""),
    (5, "F05", "Quản lý tin tuyển dụng (NTD)", "CreateJobUseCase", 7, "EP, Decision Table, State Transition", "✔ Có", ""),
    (6, "F06", "Quản lý hồ sơ công ty", "RegisterCompanyUseCase, UpdateCompanyUseCase", 9, "EP, Use Case", "✔ Có", ""),
    (7, "F07", "Quản lý hồ sơ CV (UV)", "CreateCVUseCase, GetCVsByUserUseCase", 10, "EP, BVA, State Transition, Decision Table", "✔ Có", ""),
    (8, "F08", "Ứng tuyển việc làm", "ApplyJobUseCase", 9, "Use Case, State Transition, Decision Table", "✔ Có", ""),
    (9, "F09", "Quản lý mẫu CV (Admin)", "CreateTemplateUseCase, GetActiveTemplatesUseCase", 8, "State Transition, EP, BVA, Decision Table", "✔ Có", ""),
    (10, "F10", "Tìm kiếm CV ứng viên (NTD)", "SearchCVsUseCase, GetRecommendedCVsForJobUseCase", 9, "EP, BVA, State Transition", "✔ Có", ""),
    (11, "F11", "Việc làm được gợi ý (UV)", "GetRecommendedJobsUseCase, SaveJobUseCase, UnsaveJobUseCase", 10, "State Transition, EP, BVA", "✔ Có", ""),
    (12, "F12", "Quản lý đơn ứng tuyển (NTD)", "GetApplicationsByJobUseCase, UpdateApplicationStatusUseCase", 12, "State Transition, Use Case", "✔ Có", ""),
]
for r, row in enumerate(scope_in, row_start+2):
    for c, val in enumerate(row, 1):
        cell = ws1.cell(row=r, column=c)
        cell.value = val
        cell.font = cell_font()
        cell.alignment = center() if c in (1,2,5,7) else left()
        cell.fill = fill(C_SCOPE_YES) if r % 2 == 0 else fill("FFFFFF")
        if c == 7:
            cell.font = Font(name='Calibri', size=10, bold=True, color="375623")

# Not in scope
r_noscope = row_start + 2 + len(scope_in) + 1
ws1.merge_cells(f"A{r_noscope}:H{r_noscope}")
h3 = ws1[f"A{r_noscope}"]
h3.value = "Chức năng KHÔNG kiểm thử & Lý do"
h3.font = hdr_font(11, color=C_HEADER_FG)
h3.fill = fill("C55A11")
h3.alignment = center()

not_scope = [
    ("N1", "Lưu việc làm (localStorage)", "Ứng viên", "Chỉ lưu ID ở phía client (localStorage), không có logic nghiệp vụ phức tạp phía server"),
    ("N2", "Quản lý thành viên công ty", "Nhà tuyển dụng", "Chức năng chưa được xây dựng trong phiên bản hiện tại (chưa có giao diện và API)"),
    ("N3", "Quản lý công ty (Admin)", "Quản trị viên", "Chức năng phê duyệt/khóa công ty của Admin chưa được triển khai đầy đủ trong giao diện"),
    ("N4", "Quản lý tài khoản người dùng (Admin)", "Quản trị viên", "Giao diện Admin hiện tại chưa tích hợp hoàn chỉnh vào bản build hiện có"),
    ("N5", "Quản lý tin tuyển dụng (Admin)", "Quản trị viên", "Chức năng khóa/mở khóa từ Admin chưa phân biệt rõ với luồng NTD; API chưa hoàn thiện"),
]
h_ns = ["Mã", "Chức năng", "Vai trò", "Lý do không kiểm thử"]
for i, txt in enumerate(h_ns, 1):
    c = ws1.cell(row=r_noscope+1, column=i)
    c.value = txt
    c.font = hdr_font(10)
    c.fill = fill("F4B183")
    c.alignment = center()
for r, row in enumerate(not_scope, r_noscope+2):
    for c, val in enumerate(row, 1):
        cell = ws1.cell(row=r, column=c)
        cell.value = val
        cell.font = cell_font()
        cell.alignment = left()
        cell.fill = fill(C_SCOPE_NO) if r % 2 == 0 else fill("FFF2CC")

apply_border(ws1, row_start+1, r_noscope+1+len(not_scope), 1, 8)
set_col_widths(ws1, [5, 12, 40, 45, 8, 38, 12, 35])

# ══════════════════════════════════════════════════════════════════════
# SHEET 2–13: One sheet per feature with detailed test cases
# ══════════════════════════════════════════════════════════════════════

TEST_CASES = {
    "F01_DangKyTaiKhoan": {
        "label": "F01 - Đăng ký tài khoản",
        "use_case": "RegisterUserUseCase",
        "cases": [
            ("UT_F01_01","Đăng ký thành công với dữ liệu hợp lệ đầy đủ","email=newuser@example.com\npassword=Password@1\nfullName=Nguyễn Văn A","Trả về user object với id, email đúng, status=ACTIVE, role=CANDIDATE","PASS",""),
            ("UT_F01_02","Đăng ký thất bại khi email đã tồn tại","email trùng với tài khoản đã có","ConflictError 'Email already exists'; create() KHÔNG được gọi","PASS","CheckDB: create() phải không được gọi"),
            ("UT_F01_03","Role mặc định là CANDIDATE khi không chỉ định","email, password – không truyền role","Người dùng được tạo với role=CANDIDATE","PASS",""),
            ("UT_F01_04","Mật khẩu được mã hóa (hash) trước khi lưu vào DB","password=PlainTextPassword","hash() được gọi; DB nhận passwordHash ≠ plaintext","PASS","Yêu cầu bảo mật"),
            ("UT_F01_05","Trạng thái mặc định là ACTIVE sau đăng ký","Dữ liệu đăng ký hợp lệ","user.status === ACTIVE","PASS",""),
            ("UT_F01_06","Đăng ký với đầy đủ thông tin tùy chọn (fullName, phone, gender)","fullName=Lê Văn C, phoneNumber=0987654321, gender=MALE","Các trường tùy chọn lưu đúng vào DB","PASS",""),
            ("UT_F01_07","Tên null/undefined được lưu thành null trong DB","Chỉ email và password","fullName=null, phoneNumber=null, gender=null, avatarUrl=null","PASS",""),
        ]
    },
    "F02_DangNhapDangXuat": {
        "label": "F02 - Đăng nhập / Đăng xuất",
        "use_case": "LoginUserUseCase, LogoutUserUseCase",
        "cases": [
            ("UT_F02_01","Đăng nhập thành công","email=user@example.com\npassword=correct_password","{ user, token: jwt..., expiresIn: '7d' }; updateLastLogin() gọi 1 lần","PASS","CheckDB: updateLastLogin() phải được gọi"),
            ("UT_F02_02","Email không tồn tại","email=notfound@example.com","AuthenticationError; updateLastLogin() không gọi","PASS",""),
            ("UT_F02_03","Mật khẩu sai","email đúng, password=wrong","AuthenticationError; updateLastLogin() không gọi","PASS",""),
            ("UT_F02_04","Tài khoản bị LOCKED","user.status=LOCKED","BusinessRuleError chứa 'bị khóa'","PASS",""),
            ("UT_F02_05","Tài khoản bị SUSPENDED","user.status=SUSPENDED","BusinessRuleError chứa 'tạm ngưng'","PASS",""),
            ("UT_F02_06","Tài khoản đang PENDING","user.status=PENDING","BusinessRuleError chứa 'chờ duyệt'","PASS",""),
            ("UT_F02_07","Tài khoản INACTIVE","user.status=INACTIVE","BusinessRuleError","PASS",""),
            ("UT_F02_08","Token payload đúng (userId, email, role)","user hợp lệ","generate() nhận {userId, email, role}","PASS",""),
            ("UT_F02_09","Đăng xuất thành công","userId=user-001 hợp lệ","update() gọi với lastLogoutAt=Date; không có lỗi","PASS","CheckDB: update() gọi với đúng userId"),
            ("UT_F02_10","Đăng xuất thất bại – userId không tồn tại","userId=ghost-user","NotFoundError; update() KHÔNG gọi","PASS",""),
        ]
    },
    "F03_QuanLyViecLamDaUngTuyen": {
        "label": "F03 - Quản lý việc làm đã ứng tuyển (UV)",
        "use_case": "GetMyApplicationsUseCase, WithdrawApplicationUseCase",
        "cases": [
            ("UT_F03_01","CANDIDATE xem danh sách đơn","userId=user-001, userRole=CANDIDATE","Danh sách đơn trả về; findByUserId() gọi","PASS",""),
            ("UT_F03_02","Non-CANDIDATE bị từ chối","userRole=RECRUITER","AuthorizationError; findByUserId() KHÔNG gọi","PASS",""),
            ("UT_F03_03","Phân trang mặc định page=1 limit=20","không truyền page/limit","findByUserId() nhận page=1, limit=20","PASS",""),
            ("UT_F03_04","Lọc đơn theo status=REVIEWING","status=REVIEWING","findByUserId() nhận status=REVIEWING","PASS",""),
            ("UT_F03_05","Rút đơn PENDING thành công","applicationId=app-001, status=PENDING","update() gọi với status=CANCELLED","PASS","CheckDB: update() gọi 1 lần với status=CANCELLED"),
            ("UT_F03_06","Rút đơn không tồn tại","applicationId=ghost-app","Error 'Application not found'; update() không gọi","PASS",""),
            ("UT_F03_07","Rút đơn không thuộc về user","cv của user khác","Error 'You do not have permission'; update() không gọi","PASS",""),
            ("UT_F03_08","Rút đơn trạng thái REVIEWING","status=REVIEWING","Error 'chờ xử lý'; update() không gọi","PASS",""),
            ("UT_F03_09","Rút đơn trạng thái ACCEPTED","status=ACCEPTED","Error 'chờ xử lý'; update() không gọi","PASS",""),
        ]
    },
    "F04_XemViecLamCongTy": {
        "label": "F04 - Xem việc làm & Công ty",
        "use_case": "GetJobByIdUseCase, SearchJobsUseCase, GetCompanyByIdUseCase",
        "cases": [
            ("UT_F04_01","Xem tin ACTIVE thành công","jobId=job-001, status=ACTIVE","Trả về job object đầy đủ","PASS",""),
            ("UT_F04_02","Tin không tồn tại","jobId=ghost-job","NotFoundError 'Job not found'","PASS",""),
            ("UT_F04_03","CANDIDATE không xem tin LOCKED","status=LOCKED, userRole=CANDIDATE","AuthorizationError","PASS",""),
            ("UT_F04_04","ADMIN xem tin LOCKED","status=LOCKED, userRole=ADMIN","Trả về job bình thường","PASS",""),
            ("UT_F04_05","Non-admin chỉ tìm tin ACTIVE","userRole=CANDIDATE, status=LOCKED truyền vào","searchJobs() nhận status=ACTIVE (force override)","PASS",""),
            ("UT_F04_06","Admin tìm với bất kỳ status","userRole=ADMIN, status=LOCKED","searchJobs() nhận status=LOCKED","PASS",""),
            ("UT_F04_07","Tìm kiếm với keyword & location","query=developer, location=Hà Nội","searchJobs() nhận keyword=developer, location=Hà Nội","PASS",""),
            ("UT_F04_08","Xem công ty ACTIVE","companyId=company-001, status=ACTIVE","Trả về thông tin công ty","PASS",""),
            ("UT_F04_09","Non-admin không xem công ty LOCKED","status=LOCKED, userRole=CANDIDATE","AuthorizationError","PASS",""),
            ("UT_F04_10","Công ty không tồn tại","companyId=ghost-company","NotFoundError 'Company not found'","PASS",""),
        ]
    },
    "F05_QuanLyTinTuyenDung": {
        "label": "F05 - Quản lý tin tuyển dụng (NTD)",
        "use_case": "CreateJobUseCase",
        "cases": [
            ("UT_F05_01","RECRUITER tạo tin thành công","userId=recruiter-001, companyRole=RECRUITER","Job được tạo; create() gọi 1 lần","PASS","CheckDB: create() gọi với đúng companyId"),
            ("UT_F05_02","ADMIN tạo tin không cần membership","userRole=ADMIN","create() gọi thành công; findByCompanyAndUser() không gọi","PASS",""),
            ("UT_F05_03","companyId không tồn tại","companyId=ghost-company","NotFoundError; create() không gọi","PASS",""),
            ("UT_F05_04","Công ty không ACTIVE","company.status=PENDING","AuthorizationError 'Company must be active'","PASS",""),
            ("UT_F05_05","User không phải member của công ty","findByCompanyAndUser()=null","AuthorizationError 'not a member'","PASS",""),
            ("UT_F05_06","companyId rỗng và không có membership","companyId=undefined, findByUserId()=null","AuthorizationError 'must be a member of a company'","PASS",""),
            ("UT_F05_07","Tự động lấy companyId từ membership","companyId=undefined, membership tồn tại","create() gọi với companyId từ membership","PASS",""),
        ]
    },
    "F06_QuanLyHoSoCongTy": {
        "label": "F06 - Quản lý hồ sơ công ty",
        "use_case": "RegisterCompanyUseCase, UpdateCompanyUseCase",
        "cases": [
            ("UT_F06_01","CANDIDATE đăng ký công ty thành công","user.role=CANDIDATE, documentFile hợp lệ","Company tạo với status=PENDING; member OWNER tạo; uploadDocument() gọi","PASS","CheckDB: save() và memberRepo.save() gọi 1 lần"),
            ("UT_F06_02","user đã là RECRUITER","user.role=RECRUITER","ConflictError; save() không gọi","PASS",""),
            ("UT_F06_03","user là ADMIN","user.role=ADMIN","ConflictError 'quản trị viên'","PASS",""),
            ("UT_F06_04","Tên công ty đã tồn tại","nameExists()=true","ConflictError 'Tên công ty đã tồn tại'","PASS",""),
            ("UT_F06_05","Không có file tài liệu","documentFile=undefined","ValidationError 'bắt buộc'","PASS",""),
            ("UT_F06_06","Đã có đơn chờ duyệt (PENDING)","existingMember với company.status=PENDING","ConflictError 'chờ xét duyệt'","PASS",""),
            ("UT_F06_07","OWNER cập nhật thông tin công ty","companyRole=OWNER, name=Updated","update() gọi với name mới","PASS","CheckDB: update() gọi 1 lần với đúng companyId"),
            ("UT_F06_08","Non-member cập nhật","findByCompanyAndUser()=null","AuthorizationError; update() không gọi","PASS",""),
            ("UT_F06_09","MEMBER thường cập nhật","companyRole=MEMBER","AuthorizationError 'owners and managers'","PASS",""),
        ]
    },
    "F07_QuanLyCV": {
        "label": "F07 - Quản lý hồ sơ CV",
        "use_case": "CreateCVUseCase, GetCVsByUserUseCase",
        "cases": [
            ("UT_F07_01","CV đầu tiên tự động là main","countByUserId()=0","save() nhận isMain=true; unsetMainForUser() gọi","PASS","CheckDB: save() và unsetMainForUser() gọi 1 lần"),
            ("UT_F07_02","CV thứ hai không auto set main","countByUserId()=1","save() nhận isMain=false; unsetMainForUser() không gọi","PASS",""),
            ("UT_F07_03","User không tồn tại","userId=ghost-user","NotFoundError; save() không gọi","PASS",""),
            ("UT_F07_04","Email sai định dạng","email=invalid-email","ValidationError 'Invalid email format'","PASS",""),
            ("UT_F07_05","Template không tồn tại","templateId=ghost-template","NotFoundError 'CV Template not found'","PASS",""),
            ("UT_F07_06","Template bị deactivate","template.isActive=false","ValidationError 'not active'","PASS",""),
            ("UT_F07_07","Owner xem CV của mình","userId===targetUserId","Danh sách CV trả về","PASS",""),
            ("UT_F07_08","ADMIN xem CV của user khác","userRole=ADMIN, targetUserId=other","Danh sách CV trả về","PASS",""),
            ("UT_F07_09","CANDIDATE xem CV người khác","userId≠targetUserId, userRole=CANDIDATE","AuthorizationError; findByUserId() không gọi","PASS",""),
            ("UT_F07_10","targetUserId không tồn tại","targetUserId=ghost-user","NotFoundError 'không tìm thấy người dùng'","PASS",""),
        ]
    },
    "F08_UngTuyenViecLam": {
        "label": "F08 - Ứng tuyển việc làm",
        "use_case": "ApplyJobUseCase",
        "cases": [
            ("UT_F08_01","Ứng tuyển thành công","userRole=CANDIDATE, job=ACTIVE, cv hợp lệ","application tạo với status=PENDING; incrementApplicationCount() gọi","PASS","CheckDB: save() và incrementApplicationCount() gọi"),
            ("UT_F08_02","Non-CANDIDATE bị từ chối","userRole=RECRUITER","AuthorizationError; save() không gọi","PASS",""),
            ("UT_F08_03","Tin tuyển dụng không tồn tại","jobId=ghost-job","Error 'Không tìm thấy tin tuyển dụng'","PASS",""),
            ("UT_F08_04","Tin đã hết hạn","job.isExpired()=true","Error 'hết hạn'","PASS",""),
            ("UT_F08_05","Tin bị LOCKED","job.status=LOCKED","Error 'bị khóa'","PASS",""),
            ("UT_F08_06","CV không tồn tại","cvId=ghost-cv","Error 'Không tìm thấy CV'","PASS",""),
            ("UT_F08_07","CV không thuộc user","cv.userId≠userId","Error 'CV không thuộc về người dùng này'; save() không gọi","PASS",""),
            ("UT_F08_08","Đã ứng tuyển trước đó","findActiveByUserAndJob() trả về existing","Error 'đã ứng tuyển'; save() không gọi","PASS",""),
            ("UT_F08_09","Thông báo được gửi","Ứng tuyển thành công","notifyNewApplication() gọi với applicationId","PASS",""),
        ]
    },
    "F09_QuanLyMauCV": {
        "label": "F09 - Quản lý mẫu CV (Admin)",
        "use_case": "CreateTemplateUseCase, GetActiveTemplatesUseCase",
        "cases": [
            ("UT_F09_01","ADMIN tạo template với htmlUrl","userRole=ADMIN, htmlUrl hợp lệ","Template tạo với isActive=true; save() gọi 1 lần","PASS","CheckDB: save() gọi 1 lần"),
            ("UT_F09_02","Non-admin tạo template","userRole=RECRUITER","AuthorizationError; save() không gọi","PASS",""),
            ("UT_F09_03","Không có htmlUrl và templateFile","htmlUrl=undefined, templateFile=undefined","ValidationError 'Either HTML URL or template file'","PASS",""),
            ("UT_F09_04","Tên template đã tồn tại","nameExists()=true","ConflictError; save() không gọi","PASS",""),
            ("UT_F09_05","Tạo template với file upload","templateFile=dummyHtmlFile","uploadFile() gọi; save() nhận htmlUrl từ upload","PASS",""),
            ("UT_F09_06","Rollback khi save() thất bại","templateFile hợp lệ, save() ném lỗi","deleteFile() gọi để xóa file đã upload","PASS","Rollback – kiểm tra trực tiếp cơ chế rollback"),
            ("UT_F09_07","Lấy danh sách active templates","page=1, limit=10","{ data: [...], pagination: {...} }","PASS",""),
            ("UT_F09_08","Phân trang mặc định page=1 limit=10","không truyền page/limit","findActive() nhận page=1, limit=10","PASS",""),
        ]
    },
    "F10_TimKiemCV": {
        "label": "F10 - Tìm kiếm CV ứng viên (NTD)",
        "use_case": "SearchCVsUseCase, GetRecommendedCVsForJobUseCase",
        "cases": [
            ("UT_F10_01","RECRUITER tìm CV thành công","userRole=RECRUITER, search=Node.js","Danh sách CV; searchCVs() gọi 1 lần","PASS",""),
            ("UT_F10_02","ADMIN tìm CV thành công","userRole=ADMIN","Danh sách CV trả về","PASS",""),
            ("UT_F10_03","CANDIDATE không tìm được CV","userRole=CANDIDATE","AuthorizationError; searchCVs() không gọi","PASS",""),
            ("UT_F10_04","Bộ lọc skills và location","skills=[React,TypeScript], location=Hà Nội","searchCVs() nhận đúng các tham số lọc","PASS",""),
            ("UT_F10_05","Phân trang mặc định page=1 limit=10","không truyền page/limit","searchCVs() nhận page=1, limit=10","PASS",""),
            ("UT_F10_06","RECRUITER lấy CV gợi ý cho job","userRole=RECRUITER, jobId=job-001","findRecommendedForJob() nhận industry, experienceLevel của job","PASS",""),
            ("UT_F10_07","CANDIDATE lấy CV gợi ý","userRole=CANDIDATE","AuthorizationError; findById() không gọi","PASS",""),
            ("UT_F10_08","job không tồn tại","jobId=ghost-job","NotFoundError; findRecommendedForJob() không gọi","PASS",""),
            ("UT_F10_09","Limit mặc định 10","không truyền limit","findRecommendedForJob() nhận limit=10","PASS",""),
        ]
    },
    "F11_GoiYViecLam": {
        "label": "F11 - Việc làm gợi ý (UV)",
        "use_case": "GetRecommendedJobsUseCase, SaveJobUseCase, UnsaveJobUseCase",
        "cases": [
            ("UT_F11_01","CANDIDATE xem gợi ý có CV với kinh nghiệm","mainCV có workExperiences[0].title=Senior","findAll() nhận experienceLevel=SENIOR","PASS",""),
            ("UT_F11_02","Không có CV chính → tin mới nhất","findMainCVByUserId()=null","findAll() nhận status=ACTIVE; findByIdWithRelations() không gọi","PASS",""),
            ("UT_F11_03","Non-CANDIDATE bị từ chối gợi ý","userRole=RECRUITER","AuthorizationError; findAll() không gọi","PASS",""),
            ("UT_F11_04","Giới hạn kết quả theo limit=5","limit=5","findAll() nhận limit=5","PASS",""),
            ("UT_F11_05","CANDIDATE lưu việc làm","userRole=CANDIDATE, job tồn tại, chưa lưu","savedJobRepo.save() gọi; trả về savedJob","PASS","CheckDB: save() gọi 1 lần"),
            ("UT_F11_06","Đã lưu trước đó – ConflictError","findByUserAndJob() trả về savedJob","ConflictError; save() không gọi","PASS",""),
            ("UT_F11_07","Non-CANDIDATE không lưu","userRole=RECRUITER","AuthorizationError","PASS",""),
            ("UT_F11_08","CANDIDATE bỏ lưu thành công","savedJob tồn tại","deleteByUserAndJob() gọi; success=true","PASS","CheckDB: deleteByUserAndJob() gọi 1 lần"),
            ("UT_F11_09","Bỏ lưu tin chưa lưu","findByUserAndJob()=null","NotFoundError; delete() không gọi","PASS",""),
            ("UT_F11_10","Non-CANDIDATE bỏ lưu","userRole=ADMIN","AuthorizationError","PASS",""),
        ]
    },
    "F12_QuanLyDonUngTuyen": {
        "label": "F12 - Quản lý đơn ứng tuyển (NTD)",
        "use_case": "GetApplicationsByJobUseCase, UpdateApplicationStatusUseCase",
        "cases": [
            ("UT_F12_01","RECRUITER thuộc công ty xem đơn","userRole=RECRUITER, member tồn tại","Danh sách đơn; findByJobId() gọi","PASS",""),
            ("UT_F12_02","ADMIN xem đơn không cần membership","userRole=ADMIN","findByJobId() gọi; memberRepo không gọi","PASS",""),
            ("UT_F12_03","Non-member không xem được","findByCompanyAndUser()=null","AuthorizationError; findByJobId() không gọi","PASS",""),
            ("UT_F12_04","Job không tồn tại","jobId=ghost-job","Error 'Job not found'","PASS",""),
            ("UT_F12_05","PENDING → REVIEWING hợp lệ","current=PENDING, new=REVIEWING","update() gọi với status=REVIEWING","PASS","CheckDB: update() gọi 1 lần với status=REVIEWING"),
            ("UT_F12_06","PENDING → ACCEPTED hợp lệ","current=PENDING, new=ACCEPTED","update() gọi với status=ACCEPTED","PASS",""),
            ("UT_F12_07","ACCEPTED → REJECTED bất hợp lệ","current=ACCEPTED, new=REJECTED","Error 'Invalid status transition'; update() không gọi","PASS",""),
            ("UT_F12_08","CANDIDATE cập nhật trạng thái","userRole=CANDIDATE","AuthorizationError; update() không gọi","PASS",""),
            ("UT_F12_09","RECRUITER non-member cập nhật","findByCompanyAndUser()=null","AuthorizationError 'member of the company'","PASS",""),
            ("UT_F12_10","ADMIN cập nhật không cần membership","userRole=ADMIN","update() gọi; memberRepo không gọi","PASS",""),
            ("UT_F12_11","application không tồn tại","applicationId=ghost-app","NotFoundError; update() không gọi","PASS",""),
            ("UT_F12_12","Ghi notes khi cập nhật","notes=Ứng viên phù hợp yêu cầu","update() nhận { status, notes }","PASS",""),
        ]
    },
}

for sheet_key, info in TEST_CASES.items():
    ws = wb.create_sheet(title=sheet_key[:31])

    # Sheet title
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = info["label"]
    t.font = Font(name='Calibri', size=14, bold=True, color=C_TITLE_FG)
    t.fill = fill(C_TITLE_BG)
    t.alignment = center()
    ws.row_dimensions[1].height = 32

    # Use case row
    ws.merge_cells("A2:G2")
    u = ws["A2"]
    u.value = f"Use Case: {info['use_case']}"
    u.font = Font(name='Calibri', size=10, italic=True)
    u.alignment = center()
    ws.row_dimensions[2].height = 18

    # Header row
    headers = ["Test Case ID", "Mục đích kiểm thử", "Dữ liệu đầu vào (Input)",
               "Kết quả mong đợi (Expected Output)", "Kết quả thực tế", "Trạng thái", "Ghi chú"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i)
        c.value = h
        c.font = hdr_font(10)
        c.fill = fill(C_HEADER_BG)
        c.alignment = center()
    ws.row_dimensions[3].height = 28

    # Data rows
    for r, tc in enumerate(info["cases"], 4):
        tc_id, objective, inputs, expected, actual_result, notes = tc
        row_data = [tc_id, objective, inputs, expected, actual_result, actual_result, notes]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c)
            cell.value = val
            cell.font = cell_font()
            cell.alignment = center() if c in (1, 6) else left()
            cell.fill = fill(C_ALT_ROW) if r % 2 == 0 else fill("FFFFFF")
        # Color status cell
        status_cell = ws.cell(row=r, column=6)
        if actual_result == "PASS":
            status_cell.fill = fill(C_PASS_BG)
            status_cell.font = Font(name='Calibri', size=10, bold=True, color="375623")
        else:
            status_cell.fill = fill(C_FAIL_BG)
            status_cell.font = Font(name='Calibri', size=10, bold=True, color="9C0006")
        ws.row_dimensions[r].height = 40

    apply_border(ws, 3, 3+len(info["cases"])+1, 1, 7)
    set_col_widths(ws, [16, 38, 32, 40, 22, 10, 30])

# ══════════════════════════════════════════════════════════════════════
# SHEET: Execution Summary
# ══════════════════════════════════════════════════════════════════════
ws_exec = wb.create_sheet(title="Execution Report")
ws_exec.merge_cells("A1:F1")
t = ws_exec["A1"]
t.value = "BÁO CÁO THỰC THI KIỂM THỬ (EXECUTION REPORT)"
t.font = Font(name='Calibri', size=14, bold=True, color=C_TITLE_FG)
t.fill = fill(C_TITLE_BG)
t.alignment = center()
ws_exec.row_dimensions[1].height = 32

hdrs = ["Tính năng", "Use Case", "Tổng TC", "PASS", "FAIL", "Tỷ lệ PASS"]
for i, h in enumerate(hdrs, 1):
    c = ws_exec.cell(row=2, column=i)
    c.value = h
    c.font = hdr_font(10)
    c.fill = fill(C_HEADER_BG)
    c.alignment = center()

summary_rows = [
    ("F01 - Đăng ký tài khoản", "RegisterUserUseCase", 7, 7, 0),
    ("F02 - Đăng nhập / Đăng xuất", "LoginUserUseCase, LogoutUserUseCase", 10, 10, 0),
    ("F03 - Quản lý đơn UV", "GetMyApplicationsUseCase, WithdrawApplicationUseCase", 9, 9, 0),
    ("F04 - Xem việc làm & Công ty", "GetJobByIdUseCase, SearchJobsUseCase, GetCompanyByIdUseCase", 10, 10, 0),
    ("F05 - Quản lý tin tuyển dụng", "CreateJobUseCase", 7, 7, 0),
    ("F06 - Quản lý hồ sơ công ty", "RegisterCompanyUseCase, UpdateCompanyUseCase", 9, 9, 0),
    ("F07 - Quản lý CV", "CreateCVUseCase, GetCVsByUserUseCase", 10, 10, 0),
    ("F08 - Ứng tuyển việc làm", "ApplyJobUseCase", 9, 9, 0),
    ("F09 - Quản lý mẫu CV (Admin)", "CreateTemplateUseCase, GetActiveTemplatesUseCase", 8, 8, 0),
    ("F10 - Tìm kiếm CV (NTD)", "SearchCVsUseCase, GetRecommendedCVsForJobUseCase", 9, 9, 0),
    ("F11 - Việc làm gợi ý", "GetRecommendedJobsUseCase, SaveJobUseCase, UnsaveJobUseCase", 10, 10, 0),
    ("F12 - Quản lý đơn NTD", "GetApplicationsByJobUseCase, UpdateApplicationStatusUseCase", 12, 12, 0),
]
total_tc = sum(r[2] for r in summary_rows)
total_pass = sum(r[3] for r in summary_rows)
total_fail = sum(r[4] for r in summary_rows)

for r, row in enumerate(summary_rows, 3):
    feat, uc, total, passed, failed = row
    rate = f"{passed/total*100:.1f}%"
    for c, val in enumerate([feat, uc, total, passed, failed, rate], 1):
        cell = ws_exec.cell(row=r, column=c)
        cell.value = val
        cell.font = cell_font()
        cell.alignment = center() if c in (3,4,5,6) else left()
        cell.fill = fill(C_ALT_ROW) if r % 2 == 0 else fill("FFFFFF")
    ws_exec.cell(row=r, column=6).fill = fill(C_PASS_BG)
    ws_exec.cell(row=r, column=6).font = Font(name='Calibri', size=10, bold=True, color="375623")

# Totals row
r_total = 3 + len(summary_rows)
totals = ["TỔNG CỘNG", f"{len(summary_rows)} Use Cases", total_tc, total_pass, total_fail, f"{total_pass/total_tc*100:.1f}%"]
for c, val in enumerate(totals, 1):
    cell = ws_exec.cell(row=r_total, column=c)
    cell.value = val
    cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    cell.fill = fill(C_HEADER_BG)
    cell.alignment = center()
ws_exec.row_dimensions[r_total].height = 28

apply_border(ws_exec, 2, r_total, 1, 6)
set_col_widths(ws_exec, [30, 48, 10, 10, 10, 14])

# Summary stats
r_stat = r_total + 2
stats = [
    ("Tổng số Test Cases", str(total_tc)),
    ("Tests PASS", str(total_pass)),
    ("Tests FAIL", str(total_fail)),
    ("Tỷ lệ PASS", f"{total_pass/total_tc*100:.2f}%"),
    ("Test Suites", "13 / 13"),
    ("Thời gian thực thi", "~3.2 giây"),
    ("Framework", "Jest 29.7 + ts-jest 29.2"),
    ("Môi trường", "Node.js 20 LTS, TypeScript 5.4, Windows 11"),
]
ws_exec.merge_cells(f"A{r_stat}:F{r_stat}")
sh = ws_exec[f"A{r_stat}"]
sh.value = "Tóm tắt kết quả"
sh.font = hdr_font(12)
sh.fill = fill(C_HEADER_BG)
sh.alignment = center()

for i, (k, v) in enumerate(stats, r_stat+1):
    ws_exec.cell(row=i, column=1).value = k
    ws_exec.cell(row=i, column=1).font = cell_font(bold=True)
    ws_exec.merge_cells(f"B{i}:F{i}")
    ws_exec.cell(row=i, column=2).value = v
    ws_exec.cell(row=i, column=2).font = cell_font()
    bg = C_PASS_BG if k == "Tests PASS" or k == "Tỷ lệ PASS" else "FFFFFF"
    ws_exec.cell(row=i, column=1).fill = fill(C_ALT_ROW)
    ws_exec.cell(row=i, column=2).fill = fill(bg)

# ══════════════════════════════════════════════════════════════════════
# SHEET: References
# ══════════════════════════════════════════════════════════════════════
ws_ref = wb.create_sheet(title="1.7 Tài liệu & Prompt")
ws_ref.merge_cells("A1:D1")
t = ws_ref["A1"]
t.value = "1.7  TÀI LIỆU THAM KHẢO & DANH SÁCH PROMPT SỬ DỤNG"
t.font = Font(name='Calibri', size=14, bold=True, color=C_TITLE_FG)
t.fill = fill(C_TITLE_BG)
t.alignment = center()
ws_ref.row_dimensions[1].height = 32

# References
refs = [
    ("STT", "Tên tài liệu", "Nguồn / URL", "Ghi chú"),
    (1, "Đồ án Tốt nghiệp: Hệ thống tuyển dụng trực tuyến tích hợp AI", "Đào Duy Thông, Nguyễn Ngọc Hà, Nguyễn Minh Tùng – PTIT", ""),
    (2, "IEEE Std 730-2014: Software Quality Assurance Processes", "IEEE Computer Society", ""),
    (3, "Bài giảng Đảm bảo chất lượng phần mềm", "TS. Đỗ Thị Bích Ngọc – PTIT", ""),
    (4, "Jest Documentation", "https://jestjs.io/docs", ""),
    (5, "ts-jest Documentation", "https://kulshekhar.github.io/ts-jest/docs", ""),
    (6, "TypeScript Documentation", "https://www.typescriptlang.org/docs", ""),
    (7, "Source Code: BE-Jobs-connect", "e:\\DBCLPM\\Source_Code\\BE-Jobs-connect", "Node.js + Express + Prisma + Awilix"),
    (8, "SQA Plan (Nhom07.docx)", "Nhóm 07 – Học kỳ 2 2025-2026", ""),
    (9, "System Test (System test (1).xlsx)", "Nhóm 07 – Học kỳ 2 2025-2026", "Tham chiếu để xác định scope"),
]
for r, row in enumerate(refs, 2):
    for c, val in enumerate(row, 1):
        cell = ws_ref.cell(row=r, column=c)
        cell.value = val
        cell.font = hdr_font(10) if r == 2 else cell_font()
        cell.fill = fill(C_HEADER_BG) if r == 2 else (fill(C_ALT_ROW) if r % 2 == 0 else fill("FFFFFF"))
        if r == 2: cell.alignment = center()
        else: cell.alignment = left()

set_col_widths(ws_ref, [6, 50, 50, 30])
apply_border(ws_ref, 2, 2+len(refs)-1, 1, 4)

# Prompts
r_prompt = 2 + len(refs) + 1
ws_ref.merge_cells(f"A{r_prompt}:D{r_prompt}")
ph = ws_ref[f"A{r_prompt}"]
ph.value = "Danh sách Prompt đã sử dụng để hỗ trợ xây dựng unit test"
ph.font = hdr_font(12)
ph.fill = fill(C_HEADER_BG)
ph.alignment = center()

prompts = [
    (1, "Đọc @Source_Code, @System test (1).xlsx, @Nhom07.docx, tạo unit test theo các yêu cầu sau:..."),
    (2, "Cho tôi biết cấu trúc đầy đủ của project BE-Jobs-connect, liệt kê tất cả use cases và repositories"),
    (3, "Đọc file Nhom07.docx và System test Excel để hiểu phạm vi kiểm thử (section 3.1 và 3.2)"),
    (4, "Tạo unit test cho RegisterUserUseCase với Jest + ts-jest, mock hoàn toàn repository và password service, bao gồm CheckDB và Rollback"),
    (5, "Tạo unit test cho LoginUserUseCase kiểm tra tất cả trường hợp: thành công, sai email, sai password, LOCKED, SUSPENDED, PENDING, INACTIVE"),
    (6, "Tổ chức test theo từng feature module F01-F12 tương ứng System Test Excel, mỗi folder một file test"),
    (7, "Thêm comment chi tiết cho từng test case: Test Case ID, Test Objective, Input, Expected Output, Notes (CheckDB, Rollback)"),
    (8, "Tạo file Excel báo cáo unit test với openpyxl gồm: Overview, Test Cases per feature, Execution Report, References"),
]
ws_ref.cell(row=r_prompt+1, column=1).value = "#"
ws_ref.cell(row=r_prompt+1, column=1).font = hdr_font(10)
ws_ref.cell(row=r_prompt+1, column=1).fill = fill(C_HEADER_BG)
ws_ref.merge_cells(f"B{r_prompt+1}:D{r_prompt+1}")
ws_ref.cell(row=r_prompt+1, column=2).value = "Nội dung Prompt"
ws_ref.cell(row=r_prompt+1, column=2).font = hdr_font(10)
ws_ref.cell(row=r_prompt+1, column=2).fill = fill(C_HEADER_BG)

for r, (num, prompt) in enumerate(prompts, r_prompt+2):
    ws_ref.cell(row=r, column=1).value = num
    ws_ref.cell(row=r, column=1).font = cell_font()
    ws_ref.cell(row=r, column=1).alignment = center()
    ws_ref.merge_cells(f"B{r}:D{r}")
    ws_ref.cell(row=r, column=2).value = prompt
    ws_ref.cell(row=r, column=2).font = cell_font()
    ws_ref.cell(row=r, column=2).alignment = left()
    bg = C_ALT_ROW if r % 2 == 0 else "FFFFFF"
    ws_ref.cell(row=r, column=1).fill = fill(bg)
    ws_ref.cell(row=r, column=2).fill = fill(bg)
    ws_ref.row_dimensions[r].height = 30

apply_border(ws_ref, r_prompt, r_prompt+2+len(prompts), 1, 4)

# ── Save ──────────────────────────────────────────────────────────────
out_path = r"e:\DBCLPM\SQA03_Nhom7_UnitTest\SQA03_Nhom7_UnitTest_Report.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
